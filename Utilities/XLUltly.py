import openpyxl

def getRowCount(file,sheetName):
    filename = openpyxl.load_workbook(file)
    sheet = filename[sheetName]
    return(sheet.max_row)

def getColumnCount(file,sheetName):
    filename = openpyxl.load_workbook(file)
    sheet = filename[sheetName]
    return(sheet.max_column)

def readdata(file,sheetName,rownum,colnum):
    filename = openpyxl.load_workbook(file)
    sheet = filename[sheetName]
    return sheet.cell(row=rownum, column= colnum).value

def wriredata(file,sheetName,rownum,colnum,data):
    filename = openpyxl.load_workbook(file)
    sheet = filename[sheetName]
    sheet.cell(row=rownum, column= colnum).value = data
    filename.save(file)
